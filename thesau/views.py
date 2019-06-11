from django.contrib.auth.models import User
from django.core.files import File
from django.core.files.base import ContentFile
from django.db.models import Sum
from django.shortcuts import render, redirect
from django.utils.datetime_safe import datetime

from ds4admin.utils import check_moveout_dinners
from ds4reboot.settings import TEMP_FOLDER, MEDIA_ROOT, HR_REPORTS_FOLDER
from thesau.models import Report, BoetesReport, UserReport
from user.models import Housemate
from django.contrib import messages
from openpyxl import Workbook


# view for thesau page
def index(request):
    if request.user.groups.filter(name='thesau').exists() or request.user.is_superuser:

        # get reports archive
        report_list = Report.objects.all().order_by('id')

        # build context object
        context = {
            'breadcrumbs': request.get_full_path()[1:-1].split('/'),
            'report_list': report_list,
        }
        return render(request, 'thesau/index.html', context)

    else:
        messages.error(request, 'Only accessible to thesaus and admins.')
        return redirect('/')


# view for HR page
def hr(request):
    if request.user.groups.filter(name='thesau').exists() or request.user.is_superuser:

        # generate necessary user lists
        active_users = User.objects.filter(is_active=True).exclude(username='admin')
        user_list = Housemate.objects.filter(user__id__in=active_users).order_by('movein_date')

        moveout_open_dinners, moveout_pending = check_moveout_dinners(request)

        # calculate turf totals
        totals = [str(list(user_list.aggregate(Sum('sum_bier')).values())[0]),
                  str(list(user_list.aggregate(Sum('sum_wwijn')).values())[0]),
                  str(list(user_list.aggregate(Sum('sum_rwijn')).values())[0]),
                  str(list(user_list.aggregate(Sum('boetes_geturfd_rwijn')).values())[0]),
                  str(list(user_list.aggregate(Sum('boetes_geturfd_wwijn')).values())[0])]

        # get boete counts
        # boetes_w = BoetesReport.objects.get_or_create(type='w', defaults={'boete_count': 0})
        # boetes_r = BoetesReport.objects.get_or_create(type='r', defaults={'boete_count': 0})

        # build context object
        context = {
            'breadcrumbs': request.get_full_path()[1:-1].split('/'),
            'user_list': user_list,
            'moveout_list': moveout_pending,
            'moveout_open_dinners': moveout_open_dinners,
            'boetes': [BoetesReport.objects.get(type='w').boete_count, BoetesReport.objects.get(type='r').boete_count],
            'totals': totals,
        }

        return render(request, 'thesau/hr.html', context)

    else:
        messages.error(request, 'Only accessible to thesaus and admins.')
        return redirect('/')


# generate XLSX file and commit changes
def submit_hr(request):
    # generate necessary user lists
    active_users = User.objects.filter(is_active=True).exclude(username='admin')
    user_list = Housemate.objects.filter(user__id__in=active_users).order_by('movein_date')
    moveout_list = Housemate.objects.filter(moveout_set=1).order_by('moveout_date')

    # calculate turf totals
    totals = [list(user_list.aggregate(Sum('sum_bier')).values())[0],
              list(user_list.aggregate(Sum('sum_wwijn')).values())[0],
              list(user_list.aggregate(Sum('sum_rwijn')).values())[0],
              list(user_list.aggregate(Sum('boetes_geturfd_rwijn')).values())[0],
              list(user_list.aggregate(Sum('boetes_geturfd_wwijn')).values())[0]]

    # get boete counts
    boetes_rwijn = BoetesReport.objects.get(type='r')
    boetes_wwijn = BoetesReport.objects.get(type='w')

    # make workbook and select active worksheet
    wb = Workbook()
    ws1 = wb.active

    ws1.title = 'Bierlijst'
    ws1.append(['Naam', 'Bier', 'W. Wijn', 'R. Wijn', 'Boetewijn Rood', 'Boetewijn Wit'])

    for u in user_list:
        ws1.append(
            [u.display_name, u.sum_bier, u.sum_wwijn, u.sum_rwijn, u.boetes_geturfd_rwijn, u.boetes_geturfd_wwijn])

    ws1.append([''])
    ws1.append(['Totaal', totals[0], totals[1], totals[2], totals[3], totals[4]])
    ws1.append([''])

    # Latest HR date
    latest_hr = Report.objects.latest('id')

    ws1.append(['Moved out housemates below'])
    ws1.append(['Naam', 'Bier', 'W. Wijn', 'R. Wijn', 'Boetewijn Rood', 'Boetewijn Wit', 'Open'])
    if moveout_list:
        for u in moveout_list:
            if u.moveout_date >= latest_hr.report_date:
                ws1.append([u.display_name, u.sum_bier, u.sum_wwijn, u.sum_rwijn, u.boetes_geturfd_rwijn,
                            u.boetes_geturfd_wwijn, u.boetes_open])

    # create secondary worksheets
    ws2 = wb.create_sheet()
    ws2.title = 'Geturfd Boetes'

    ws2.append(['W. Wijn', 'R. Wijn'])
    ws2.append([boetes_wwijn.boete_count, boetes_rwijn.boete_count])

    # reset boetes count
    boetes_wwijn.boete_count = 0
    boetes_rwijn.boete_count = 0

    if moveout_list:
        ws3 = wb.create_sheet()

        ws3.title = 'Oudhuisgenoten'
        ws3.sheet_properties.tabColor = "FB29B4"
        ws3.append(['Naam',
                    'Verhuizing datum', 'Laatste HR datum',
                    'Huidige HR datum',
                    'Dagen tot laatste HR',
                    'Dagen tot huidige HR',
                    'Percentage HR'
                    'Gecompenseerd saldo'])

        latest_hr_date = latest_hr.report_date
        now_date = datetime.now().date()

        for u in moveout_list:

            if u.moveout_date >= latest_hr.report_date:

                days_latest_hr = (u.moveout_date - latest_hr_date).days
                days_current_hr = (now_date - u.moveout_date).days

                if days_current_hr == 0 and days_latest_hr == 0:
                    perc = 0
                else:
                    perc = float(days_latest_hr / (days_latest_hr + days_current_hr))

                ws3.append([u.display_name,
                            u.moveout_date, latest_hr_date,
                            now_date,
                            days_latest_hr,
                            days_current_hr,
                            perc,
                            u.balance])

                # This action comes from remove_housemate in ds4admin page
                u.balance = 0.00

    # save the file
    date = datetime.now()
    months = {1: 'januari', 2: 'februari', 3: 'maart', 4: 'april', 5: 'mei', 6: 'juni', 7: 'juli',
              8: 'augustus', 9: 'september', 10: 'oktober', 11: 'november', 12: 'december'}

    temp_path = MEDIA_ROOT + TEMP_FOLDER + f"HR_{date.year}_{months[date.month]}.xlsx"

    wb.save(temp_path)
    file_ref = open(temp_path, 'rb')
    file_content = File(file_ref)
    # create report entry in database
    r = Report(report_user=request.user,
               report_name='HR %s-%s-%s (%s:%s)' % (date.day, months[date.month], date.year, date.hour, date.minute),
               report_date=date)
    r.report_file.save(f"HR_{date.year}_{months[date.month]}.xlsx", file_content)
    r.save()
    file_ref.close()

    # Save users, workbook, boetes and report
    report_users = user_list | moveout_list
    for u in report_users:
        if u in moveout_list:
            open_balance = u.balance
            u.moveout_set = True
        else:
            open_balance = 0

        ur = UserReport(user=u.user, report=r,
                        hr_bier=u.sum_bier, hr_wwijn=u.sum_wwijn, hr_rwijn=u.sum_rwijn,
                        hr_boete_rwijn=u.boetes_geturfd_rwijn, hr_boete_wwijn=u.boetes_geturfd_wwijn,
                        eetlijst_balance=open_balance)
        ur.save()

        u.sum_bier = 0
        u.sum_wwijn = 0
        u.sum_rwijn = 0
        u.boetes_geturfd_rwijn = 0
        u.boetes_geturfd_wwijn = 0

    ##
    # SAVE AT THE END
    ##
    for u in report_users:
        u.save()

    for u in moveout_list:
        u.save()

    boetes_wwijn.save()
    boetes_rwijn.save()

    return redirect('/thesau/')
