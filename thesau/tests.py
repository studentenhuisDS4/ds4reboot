import os
import io
import pytz

from django.contrib.auth.models import User
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils.datetime_safe import date, datetime
from django.conf.global_settings import DEBUG
from django.utils.timezone import make_aware
from django.utils import timezone
from rainbowtests import colors
from openpyxl.reader.excel import load_workbook

from ds4reboot.tests import assert_total_balance
from eetlijst.models import Dinner, UserDinner
from thesau.models import Report, BoetesReport
from thesau.urls import urlpatterns
from user.models import Housemate
from bierlijst.models import Boete


class ThesauTest(TestCase):
    SUPER_PASS = 'Studentenhuis'

    skip_nonstatic = ['hr report', 'submit hr']
    dinner_enroll_url = reverse('enroll')
    dinner_close_url = reverse('close')
    dinner_cost_url = reverse('cost')
    remove_housemate_url = reverse('remove housemate')
    hr_url = reverse('submit hr')

    def setUp(self):
        print("[Testing THESAU]:")
        self.superuser = User.objects.create_superuser(
            'huis', 'studentenhuisds4@gmail.com', self.SUPER_PASS)
        Housemate.objects.create(user=User.objects.get(
            username='huis'), display_name='Huis')

        self.user = User.objects.create_superuser(
            'pietje', 'studentenhuisds4@gmail.com', 'PietjePuk')
        self.user2 = User.objects.create_superuser(
            'pietje2', 'studentenhuisds5@gmail.com', 'Pietje2Puk2')
        self.user3 = User.objects.create_superuser(
            'pietje3', 'studentenhuisds6@gmail.com', 'Pietje3Puk3')

        self.hm_pietje = Housemate.objects.create(user=User.objects.get(username='pietje'), display_name='PietjePuk',
                                                  balance=-2)
        self.hm_pietje2 = Housemate.objects.create(user=User.objects.get(username='pietje2'),
                                                   display_name='Pietje2Puk2', balance=1)
        self.hm_pietje3 = Housemate.objects.create(user=User.objects.get(username='pietje3'),
                                                   display_name='Pietje3Puk3', balance=1)

        self.assertIsNotNone(self.dinner_close_url)
        self.assertIsNotNone(self.dinner_cost_url)

        self.report = Report.objects.create(
            report_user=self.user,
            report_name="TEST REPORT",
            report_date=make_aware(datetime(2019, 4, 13), timezone=pytz.UTC))
        self.bwijn_w = BoetesReport.objects.create(type='w')
        self.bwijn_r = BoetesReport.objects.create(type='r')

    def tearDown(self):
        print(">> Teardown, removing at most {} .xlsx report files created in this test.".format(
            Report.objects.all().count()))
        for report in Report.objects.all():
            try:
                print("> Removing ", report.report_file.path)
                # os.remove(report.report_file.path)
            except ValueError:
                pass

        User.objects.filter(username='pietje').delete()
        User.objects.filter(username='pietje2').delete()
        User.objects.filter(username='pietje3').delete()
        User.objects.filter(username='huis').delete()
        Dinner.objects.all().delete()
        UserDinner.objects.all().delete()

        BoetesReport.objects.all().delete()
        Report.objects.all().delete()

    def test_static_responses(self):
        for url in urlpatterns:
            if url.name not in self.skip_nonstatic and url.name is not None:
                response = self.client.get(reverse(url.name), follow=True)
                self.assertEqual(response.status_code, 200)

    def test_boete_aggregate_regression(self):
        # Setup and Act
        print("> Regression, boete model aggregate function.")
        previous_report_date = self.report.report_date
        boete_amount = 5

        Boete.objects.all().delete()
        with override_settings(USE_TZ=False):
            b = Boete(boete_user=self.user, boete_name=self.user.housemate.display_name, created_by=self.user, boete_count=boete_amount,
                      boete_note="Regressive test Random boete.")
            b.save()
            user_boetes = Boete.aggregate_user_fines(
                latest_date=previous_report_date)

            # Assert
            self.assertEqual(user_boetes.count(), 1)
            self.assertIn('boete_sum', user_boetes[0])
            self.assertEqual(user_boetes[0]['boete_sum'], 5)

            # Cleanup
            b.delete()

    def test_boete_aggregate_previous_hr(self):
        # Setup and Act
        print("> Regression, boete model aggregate function.")
        previous_report_date = self.report.report_date
        boete_amount = 5

        Boete.objects.all().delete()
        with override_settings(USE_TZ=False):
            b = Boete(boete_user=self.user,
                      created_time=previous_report_date -
                      timezone.timedelta(1),
                      boete_name=self.user.housemate.display_name,
                      created_by=self.user,
                      boete_count=boete_amount,
                      boete_note="Old fine, should not be exported.")
            b2 = Boete(boete_user=self.user,
                       created_time=previous_report_date +
                       timezone.timedelta(1),
                       boete_name=self.user.housemate.display_name,
                       created_by=self.user,
                       boete_count=boete_amount * 2,
                       boete_note="New fine, should be exported.")
            b.save()
            b2.save()
            user_boetes = Boete.aggregate_user_fines(
                latest_date=previous_report_date)

            # Assert
            self.assertEqual(user_boetes.count(), 1)
            self.assertIn('boete_sum', user_boetes[0])
            self.assertEqual(user_boetes[0]['boete_sum'], boete_amount * 2)

            # Cleanup
            b.delete()

    def test_empty_hr_submit(self):
        print("> Testing empty hr submit")
        logged_in = self.client.login(
            username=self.superuser.username, password=self.SUPER_PASS)
        self.client.post(reverse('submit hr'))

    def test_empty_hr_export(self):
        # Quick run (within docker, otherwise python -m):
        # python ./manage.py test thesau.tests.ThesauTest.test_empty_hr_export --failfast

        print("> Testing excel export")
        boete_amount = 44
        newboete = Boete(boete_user=self.user,
                         created_time=self.report.report_date,
                         boete_name=self.user.housemate.display_name,
                         created_by=self.user.username,
                         boete_count=boete_amount,
                         boete_note="Old fine, should not be exported.")
        newboete.save()

        # Login and submit HR
        logged_in = self.client.login(
            username=self.superuser.username, password=self.SUPER_PASS)
        self.client.post(reverse('submit hr'))

        # Fetch latest report from database
        created_report = Report.objects.latest('report_date')
        file_url = created_report.report_file.path

        # Assert file and contents
        self.assertIn('.xlsx', file_url)
        wb = load_workbook(file_url)
        ws_bierlijst = wb['Bierlijst']
        cell_name = ws_bierlijst.cell(row=3, column=1).value
        cell_boetes = ws_bierlijst.cell(row=3, column=7).value
        self.assertEqual(cell_name, self.user.housemate.display_name)
        self.assertEqual(cell_boetes, boete_amount)

        wb.close()

    def test_empty_hr_submit_with_moveout(self):
        print("> Testing empty hr submit with moveout")

        logged_in = self.client.login(
            username=self.superuser.username, password=self.SUPER_PASS)

        # Moveout-set user
        user = User.objects.create_superuser(
            'normal_moveout_user', 'mail@gmail.com', 'normal_moveout_user')
        housemate = Housemate.objects.create(user=User.objects.get(username='normal_moveout_user'),
                                             display_name='NormalMoveoutUser',
                                             balance=0)
        housemate.moveout_set = True
        housemate.moveout_date = timezone.now()
        user.is_active = False
        housemate.save()
        user.save()

        response = self.client.post(reverse('submit hr'))
        self.assertEqual(response.status_code, 302)
        reports = Report.objects.count()
        self.assertEqual(reports, 2)

        latest_report = Report.objects.order_by('-id')[0]

    def test_get_presubmit_page_admin(self):
        print("> Testing hr presubmit page as admin")

        logged_in = self.client.login(
            username=self.superuser.username, password=self.SUPER_PASS)
        response = self.client.get(reverse('hr report index'))

        self.assertEqual(response.status_code, 200)

    def test_get_presubmit_page(self):
        print("> Testing hr presubmit page as normal user")

        User.objects.create_user(
            'normaluser', 'studentenhuisds4@gmail.com', 'NormalUser')
        logged_in = self.client.login(
            username='normaluser', password='NormalUser')
        response = self.client.get(reverse('hr report index'))

        self.assertEqual(response.status_code, 302)  # 302: redirect to '/'

    def test_closing_hr(self):
        print("> Testing closing HR with old housemate on dinner day")

        assert_total_balance()
        week_ago = datetime.now() - timezone.timedelta(days=7)

        # add users (normal, cook) to dinner entries
        dinner = Dinner.objects.create(date=week_ago)
        dinnerpietje = UserDinner.objects.create(
            dinner_date=week_ago, user=self.user, dinner=dinner)
        dinnerpietje.is_cook = True
        dinnerpietje.save()

        dinnerpietje2 = UserDinner.objects.create(
            dinner_date=week_ago, user=self.user2, dinner=dinner)
        dinnerpietje2.count = 1
        dinnerpietje2.save()

        # add cook to date
        dinner.cook = self.user
        dinner.num_eating = 2
        dinner.save()

        logged_in = self.client.login(username='pietje', password='PietjePuk')
        print(colors.yellow("Loggin in user for authenticated parts: "),
              colors.blue(logged_in))
        self.assertTrue(logged_in)

        print(colors.yellow("Closing dinner date: "),
              colors.blue(datetime.now().date()))
        data = {
            'close-date': week_ago.date(),
        }
        response = self.client.post(
            self.dinner_close_url, data, follow=True, HTTP_REFERER='/eetlijst/')
        self.assertIsNotNone(response)
        self.assertEqual(response.status_code, 200)

        dinner.refresh_from_db()
        self.assertFalse(dinner.open, "The dinner date was not closed!")
        print(colors.yellow("Dinner closed: "),
              colors.blue(str(not dinner.open)))

        self.hm_pietje2.refresh_from_db()
        data2 = {
            'housemate': self.hm_pietje2.user_id,
        }
        response = self.client.post(
            self.remove_housemate_url, data2, follow=True, HTTP_REFERER='/admin/')
        self.assertIsNotNone(response)
        self.assertEqual(response.status_code, 200)
        self.hm_pietje2.refresh_from_db()
        self.assertIsNone(self.hm_pietje2.moveout_date)
        print(colors.blue("Housemate pietje2 not removed because of open dinners (move-out date): "),
              colors.blue(str(self.hm_pietje2.moveout_date)))

        data3 = {
            'cost-date': week_ago.date(),
            'cost-amount': 15
        }
        response = self.client.post(
            self.dinner_cost_url, data3, follow=True, HTTP_REFERER='/eetlijst/')
        print(colors.yellow("Filled in costs."))
        assert_total_balance()

        print(colors.yellow("Submitting HR"))
        dir = './media/temp/'
        if not os.path.exists(dir):
            os.makedirs(dir)
        dir = './media/hr_reports/'
        if not os.path.exists(dir):
            os.makedirs(dir)
        response = self.client.post(
            self.hr_url, follow=True, HTTP_REFERER='/admin/')
        assert_total_balance()

        data2 = {
            'housemate': self.hm_pietje2.user_id,
        }
        response = self.client.post(
            self.remove_housemate_url, data2, follow=True, HTTP_REFERER='/admin/')
        self.assertIsNotNone(response)
        self.assertEqual(response.status_code, 200)
        self.hm_pietje2.refresh_from_db()
        self.assertIsNotNone(self.hm_pietje2.moveout_date)
        print(colors.blue("Housemate pietje2 now correctly removed (move-out date): "),
              colors.blue(str(self.hm_pietje2.moveout_date)))
        assert_total_balance()
